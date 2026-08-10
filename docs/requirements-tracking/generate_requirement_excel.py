from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT_DIR = Path(__file__).resolve().parent
OUT_FILE = OUT_DIR / "需求完成度跟踪表.xlsx"

STATUS_FILL = {
    "未开始": "F3F4F6",
    "进行中": "FEF3C7",
    "已阻塞": "FEE2E2",
    "已完成": "DCFCE7",
    "本期不做": "E5E7EB",
}

# Confirmed comments are kept here as a fallback. Existing workbook comments win.
USER_NOTES = {
    "REQ-002": "需要参考同类型其他项目",
    "REQ-006": "支持",
    "REQ-008": "可选",
    "REQ-009": "不需要必填，可添加附件",
    "REQ-010": "保留隐藏",
    "TECH-001": "优先级先",
    "TECH-003": "不需要",
    "TECH-007": "做留言机制",
    "TEST-002": "可以看",
    "TEST-003": "不需要填写测试note",
    "TEST-004": "可留填写框，但是不需要做必填",
    "TEST-005": "不新增Test Failed",
    "TEST-006": "一条",
    "TEST-007": "不限制",
    "TL-002": "优先展示",
    "TL-003": "允许",
    "TL-004": "不需要",
    "TL-005": "不需要强制备注",
    "TL-006": "需要拆分权限",
}

ROLE_SUMMARY = [
    ["Admin", "进行中", 0.96, "管理页面和请求工作流均接真实 API", "Docker/Flyway 运行验证"],
    ["普通用户", "进行中", 0.96, "创建、富文本、附件、留言和用户确认均接真实 API", "Docker 附件回归；他人详情范围待确认"],
    ["技术处理人员", "进行中", 0.95, "My Tasks、队列、分配、状态、留言和历史均接真实 API", "Docker 回归；Team Queue scope 待确认"],
    ["测试人员", "进行中", 0.94, "Test Queue、内部备注和测试状态均接真实 API", "Docker 回归"],
    ["技术负责人", "进行中", 0.95, "Assignment、队列、工作量和详情动作均接真实 API", "Docker 回归；多团队 scope 待确认"],
]

GLOBAL_PROGRESS = [
    ["产品范围", "已完成", 1.00, "旧 FAQ/服务台/Problem/Change 保留隐藏，产品页面全英文", "无"],
    ["真实 API 接入", "已完成", 1.00, "Admin、Portal、Workspace 不再引用业务 Mock", "无"],
    ["请求工作流", "进行中", 0.97, "创建、路由、分配、状态、留言、历史、富文本和附件已实现", "Docker 端到端验证"],
    ["权限与审计", "进行中", 0.95, "请求/评论/附件按角色和组织校验，动作写入 history", "他人详情粒度和团队 scope 待确认"],
    ["构建与测试", "已完成", 1.00, "前端 build、后端 compile/test 通过；5 tests, 0 failures", "无"],
    ["Docker / 本地运行", "已阻塞", 0.80, "Compose 已增加附件持久卷", "Docker daemon 未运行；V1.0.24/V1.0.25 未实库验证"],
]

COMPLETED_ITEMS = [
    ["2026-08-09", "完成登录、Portal、IT Workspace 和 Admin Console 页面", "产品可见页面全英文"],
    ["2026-08-09", "完成请求详情单一 Save 与未保存离开提示", "状态和 Assignee 统一保存"],
    ["2026-08-09", "核心请求流程接入真实 API", "登录、创建、列表、详情、分配、状态、留言和历史"],
    ["2026-08-10", "完成富文本、粘贴图片和请求/留言附件", "V1.0.24、Attachment API、服务端 HTML 清洗"],
    ["2026-08-10", "Admin 管理页面全部切换真实 API", "组织、用户、角色、权限、团队、路由、审计和 Dashboard"],
    ["2026-08-10", "完成工作流团队成员模型", "V1.0.25、Team DTO、User teamNames"],
    ["2026-08-10", "删除业务 Mock 数据源和引用", "mockAdmin.ts、mockRequests.ts 已删除"],
    ["2026-08-10", "补齐工作台排序和 Portal 状态选项约束", "优先级/未分配排序；Requester 仅确认 Resolved"],
    ["2026-08-10", "增加 Docker 附件持久卷", "request_uploads:/app/uploads"],
]

PENDING_CONFIRMATIONS = [
    ["CONF-001", "普通用户查看他人请求的详情字段和公开留言范围", "当前按组织范围控制访问，上线前确认字段粒度"],
    ["CONF-002", "Team Queue 是全 IT 还是仅团队可见", "当前 IT 范围较开放，上线前确认 scope"],
    ["CONF-003", "Team Lead 是否可负责多个团队", "模型保留扩展空间，业务规则待定"],
    ["CONF-004", "Closed 是否允许重新打开", "当前为终态，建议 P2"],
    ["CONF-005", "用户是否需要物理删除", "当前启用/禁用，建议禁用优先"],
    ["CONF-006", "Audit Logs 是否需要导出", "建议 P1"],
    ["CONF-007", "Docker 运行验证", "启动 Docker Desktop 后重建并执行多角色端到端验证"],
]


def requirement(req_id, description, priority, progress, evidence, issue="无", pending="无"):
    status = "已完成" if progress >= 1 else "进行中"
    return [req_id, description, priority, status, progress, evidence, issue, pending, ""]


ROLE_DATA = {
    "Admin": {
        "desc": "Admin 管理组织、账号、权限、团队、路由、全局请求和审计，也可作为全局技术负责人。",
        "summary": [
            ["登录与导航", "进行中", 0.97, "真实 JWT、/auth/me 和角色导航", "Docker 回归"],
            ["组织/用户/角色/权限", "进行中", 0.96, "页面和 system API 使用真实 CRUD", "Docker 写操作回归"],
            ["团队与路由", "进行中", 0.95, "真实成员、负责人、路由规则和自动匹配", "V1.0.25 实库验证"],
            ["Requests / Audit / Dashboard", "进行中", 0.96, "均使用真实 API", "Docker 数据回归"],
        ],
        "requirements": [
            requirement("ADM-001", "登录并进入 Admin Console / IT Workspace", "P0", 0.97, "admin/admin123、JWT、/auth/me", "Docker daemon 未运行"),
            requirement("ADM-002", "管理组织", "P0", 0.96, "OrganizationList.vue + Organization CRUD API", "Docker 待验证", "组织层级深度"),
            requirement("ADM-003", "管理用户和启用/禁用", "P0", 0.96, "UserList.vue + User CRUD API", "Docker 待验证", "是否增加物理删除"),
            requirement("ADM-004", "管理角色和权限", "P0", 0.96, "RoleList.vue、PermissionList.vue + 真实 API", "Docker 待验证"),
            requirement("ADM-005", "管理团队、负责人和成员", "P0", 0.95, "TeamList.vue、V1.0.25、Team/User DTO", "Flyway 未实库验证", "多团队负责人 scope"),
            requirement("ADM-006", "配置请求路由规则", "P0", 0.96, "RoutingRules.vue + RoutingRule API/匹配服务", "Docker 待验证"),
            requirement("ADM-007", "查看、分配并调整全部请求", "P0", 0.97, "Global Requests、Request Detail、单一 Save", "Docker 待验证"),
            requirement("ADM-008", "查看审计日志", "P0", 0.95, "AuditLogList.vue + Audit API", "Docker 待验证", "导出是否列为 P1"),
            requirement("ADM-009", "Admin 页面全部英文", "P0", 1.00, "新 Admin/Workspace 可见页面"),
            requirement("ADM-010", "不显示旧 FAQ/服务台/Problem/Known Error/Change", "P0", 1.00, "新菜单和路由不暴露；旧文件保留隐藏"),
        ],
        "pages": [["Admin Overview", "是", "进行中", "Dashboard API"], ["Organizations / Users", "是", "进行中", "真实 CRUD API"], ["Roles / Permissions / Teams", "是", "进行中", "真实 CRUD API"], ["Routing Rules", "是", "进行中", "真实 Routing Rule API"], ["Requests / Audit Logs", "是", "进行中", "真实 Request/Audit API"]],
    },
    "普通用户": {
        "desc": "普通用户创建请求、公开留言，并在 Resolved 后关闭或标记 User Test Failed。",
        "summary": [["Portal 与登录", "进行中", 0.97, "真实账号和请求数据", "Docker 回归"], ["新建请求", "进行中", 0.97, "真实富文本、粘贴图片和附件", "Docker 上传回归"], ["详情与留言", "进行中", 0.96, "真实详情、评论、附件和历史", "他人详情粒度待定"], ["用户确认", "进行中", 0.97, "仅 Resolved 可 Closed/User Test Failed", "Docker 回归"]],
        "requirements": [
            requirement("REQ-001", "登录并进入 Portal", "P0", 0.97, "requester01/admin123、JWT、/auth/me", "Docker 待验证"),
            requirement("REQ-002", "创建轻量但信息完整的新请求", "P0", 0.97, "真实 create API、富文本、时间字段、粘贴图片和附件", "Docker 上传待验证"),
            requirement("REQ-003", "新请求从 New 开始或按路由成为 Assigned", "P0", 0.96, "RequestService + RoutingRuleService", "Docker 自动路由待验证"),
            requirement("REQ-004", "查看自己的请求", "P0", 0.97, "Portal 真实分页列表", "Docker 待验证"),
            requirement("REQ-005", "只读查看其他用户请求", "P0", 0.88, "后端组织范围访问控制", "详情字段粒度未定", "描述/公开留言可见范围"),
            requirement("REQ-006", "添加公开留言和附件", "P0", 0.97, "Comment API + Attachment API", "Docker 上传下载待验证"),
            requirement("REQ-007", "不能看到内部备注", "P0", 0.96, "后端过滤 internal comments/history", "Docker 权限待验证"),
            requirement("REQ-008", "从 Resolved 关闭请求", "P0", 0.97, "Portal 下拉和后端双重限制", "Docker 待验证"),
            requirement("REQ-009", "从 Resolved 标记 User Test Failed", "P0", 0.97, "Portal 下拉、后端校验、留言附件", "Docker 待验证"),
            requirement("REQ-010", "Portal 不显示 FAQ / Knowledge Base", "P0", 1.00, "路由和导航不暴露；旧文件保留隐藏"),
        ],
        "pages": [["Home", "是", "进行中", "真实请求指标"], ["New Request", "是", "进行中", "富文本和附件真实上传"], ["Request Lists", "是", "进行中", "真实分页查询"], ["Request Detail", "是", "进行中", "留言、附件、历史和确认"], ["FAQ", "否", "本期不做", "保留隐藏"]],
    },
    "技术处理人员": {
        "desc": "技术人员处理、留言、转派请求，并把修复完成的请求交给测试组。",
        "summary": [["My Tasks", "进行中", 0.96, "真实查询并按优先级排序", "Docker 回归"], ["Team Queue", "进行中", 0.90, "真实 request/team API", "可见 scope 待定"], ["状态与转派", "进行中", 0.97, "任意状态、跨团队人员、单一 Save", "Docker 回归"], ["留言与历史", "进行中", 0.96, "public/internal comment、附件、history", "Docker 权限回归"]],
        "requirements": [
            requirement("TECH-001", "查看自己的任务并优先处理高优先级", "P0", 0.96, "MyTasks.vue 按优先级和更新时间排序", "Docker 待验证"),
            requirement("TECH-002", "查看团队可见请求", "P0", 0.90, "TeamQueue.vue + 真实 API", "scope 未定", "全 IT 或团队内可见"),
            requirement("TECH-003", "通过下拉选择任意状态", "P0", 0.97, "全状态下拉和统一 Save", "Docker 待验证"),
            requirement("TECH-004", "跨团队转派给其他 IT 人员", "P0", 0.97, "Assignee 显示 team/org + assign API", "Docker 待验证"),
            requirement("TECH-005", "修复完成后转入 Testing", "P0", 0.97, "transition API + history", "Docker 待验证"),
            requirement("TECH-006", "添加内部工作备注和附件", "P0", 0.96, "internal comment + attachment API", "Docker 权限待验证"),
            requirement("TECH-007", "添加面向用户的公开留言和附件", "P1", 0.96, "public comment + attachment API", "通知未做"),
            requirement("TECH-008", "User Test Failed 后转回 In Progress", "P0", 0.97, "真实 transition API", "Docker 待验证"),
            requirement("TECH-009", "请求变更同步历史", "P0", 0.96, "创建、分配、状态、留言和附件 history", "Docker 数据待验证"),
        ],
        "pages": [["IT Workspace", "是", "进行中", "真实权限和数据"], ["My Tasks", "是", "进行中", "优先级排序"], ["Team Queue", "是", "进行中", "真实队列"], ["Request Detail", "是", "进行中", "状态、分配、留言、附件、历史"]],
    },
    "测试人员": {
        "desc": "测试人员执行内部复测；通过转 Resolved，失败转 In Progress。",
        "summary": [["Test Queue", "进行中", 0.95, "真实 Testing 请求", "Docker 回归"], ["测试结果", "进行中", 0.96, "状态下拉和单条 history", "Docker 回归"], ["备注可见性", "进行中", 0.95, "可见全部内部备注，可选留言附件", "Docker 权限回归"]],
        "requirements": [
            requirement("TEST-001", "查看 Testing 请求", "P0", 0.95, "TestQueue.vue + Request API", "Docker 待验证"),
            requirement("TEST-002", "打开详情并查看全部内部备注", "P0", 0.96, "Request Detail + IT comment scope", "Docker 待验证"),
            requirement("TEST-003", "测试通过后改为 Resolved", "P0", 0.96, "transition API，note 可选", "Docker 待验证"),
            requirement("TEST-004", "测试失败后改回 In Progress", "P0", 0.96, "transition API，可选留言/附件", "Docker 待验证"),
            requirement("TEST-005", "不新增 Test Failed 状态", "P0", 1.00, "前后端状态枚举"),
            requirement("TEST-006", "测试结果写一条历史", "P0", 0.96, "一条 STATUS_CHANGED", "Docker 数据待验证"),
            requirement("TEST-007", "状态选择不限制跳步", "P0", 0.97, "统一全状态下拉", "Docker 待验证"),
        ],
        "pages": [["Test Queue", "是", "进行中", "真实 Testing 队列"], ["Request Detail", "是", "进行中", "内部备注、附件、状态和历史"], ["Test Result Dialog", "否", "本期不做", "不强制测试备注"]],
    },
    "技术负责人": {
        "desc": "技术负责人管理团队队列、请求分配、状态和工作量；Tester 权限独立。",
        "summary": [["团队与队列", "进行中", 0.94, "真实 Team/User DTO 和 Team Queue", "多团队 scope 待定"], ["Assignment Desk", "进行中", 0.97, "未分配置顶并按优先级排序", "Docker 回归"], ["状态与分配", "进行中", 0.97, "跨团队、任意状态、单一 Save", "Docker 回归"], ["Workload", "进行中", 0.95, "真实请求统计", "专用统计接口可后续优化"]],
        "requirements": [
            requirement("TL-001", "查看团队队列", "P0", 0.94, "TeamQueue.vue + 真实 API", "Docker 待验证", "多团队 scope"),
            requirement("TL-002", "分配新请求并优先展示未分配项", "P0", 0.97, "Assignment.vue 未分配置顶、优先级排序", "Docker 待验证"),
            requirement("TL-003", "跨团队重新分配/转交", "P0", 0.97, "真实 assign API，人员显示 team/org", "Docker 待验证"),
            requirement("TL-004", "选择任意请求状态", "P0", 0.97, "全状态下拉和统一 Save", "Docker 待验证"),
            requirement("TL-005", "代技术人员提交 Testing", "P0", 0.97, "transition API，备注可选", "Docker 待验证"),
            requirement("TL-006", "Tester 权限独立分配", "P1", 0.95, "Team Lead / Tester 独立角色权限", "Docker 权限待验证"),
            requirement("TL-007", "查看团队工作量", "P1", 0.95, "Overview 基于真实请求", "可增加专用统计 API"),
            requirement("TL-008", "负责人动作写入历史", "P0", 0.96, "分配、状态、留言和附件 history", "Docker 数据待验证"),
        ],
        "pages": [["Team Queue", "是", "进行中", "真实队列"], ["Assignment Desk", "是", "进行中", "未分配和优先级排序"], ["Workload Overview", "是", "进行中", "真实请求统计"], ["Request Detail", "是", "进行中", "状态、分配和历史"]],
    },
}


def load_existing_notes():
    notes = dict(USER_NOTES)
    if not OUT_FILE.exists():
        return notes
    workbook = load_workbook(OUT_FILE, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            headers = {}
            for row in sheet.iter_rows():
                values = [cell.value for cell in row]
                if "需求ID" in values and "备注" in values:
                    headers = {value: index for index, value in enumerate(values) if value}
                    continue
                if headers:
                    request_id = values[headers["需求ID"]] if len(values) > headers["需求ID"] else None
                    note = values[headers["备注"]] if len(values) > headers["备注"] else None
                    if request_id and note:
                        notes[str(request_id)] = str(note)
    finally:
        workbook.close()
    return notes


def setup_sheet(ws):
    widths = {"A": 16, "B": 42, "C": 12, "D": 16, "E": 12, "F": 38, "G": 38, "H": 38, "I": 22}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def add_title(ws, title, description):
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws["A1"].fill = PatternFill("solid", fgColor="E5E7EB")
    ws.merge_cells("A2:I2")
    ws["A2"] = description
    ws["A2"].font = Font(size=10, color="475467")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36


def add_section(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1C2430")
    return row + 1


def write_table(ws, row, headers, values, name):
    start = row
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=column, value=header)
        cell.font = Font(bold=True, color="111827")
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    row += 1
    for item in values:
        for column, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(value, float):
                cell.number_format = "0%"
            if value in STATUS_FILL:
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL[value])
        row += 1
    end_column = len(headers)
    table = Table(displayName=name, ref=f"A{start}:{chr(64 + end_column)}{row - 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return row + 2


def style_workbook(workbook):
    thin = Side(style="thin", color="E5E7EB")
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                font = copy(cell.font)
                font.name = "Microsoft YaHei"
                cell.font = font
        validation = DataValidation(type="list", formula1='"未开始,进行中,已阻塞,已完成,本期不做"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add("B1:B500")
        validation.add("D1:D500")


def build_workbook():
    notes = load_existing_notes()
    for role in ROLE_DATA.values():
        for item in role["requirements"]:
            item[8] = notes.get(item[0], "")

    workbook = Workbook()
    ws = workbook.active
    ws.title = "总览"
    setup_sheet(ws)
    add_title(ws, "需求完成度跟踪总览", "按角色记录需求、完成度、问题和待说明事项。同步至 2026-08-10 真实 API、附件和 Admin CRUD 版本。")
    row = 4
    row = add_section(ws, row, "状态说明")
    row = write_table(ws, row, ["状态", "含义"], [["未开始", "尚未实现"], ["进行中", "主要能力已实现，仍有验证或未决范围"], ["已阻塞", "受环境或决策阻塞"], ["已完成", "已实现并完成当前阶段验证"], ["本期不做", "当前阶段明确排除"]], "StatusLegend")
    row = add_section(ws, row, "角色完成度总览")
    role_start = row
    row = write_table(ws, row, ["角色", "状态", "主观完成度", "当前证据", "主要缺口"], ROLE_SUMMARY, "RoleSummary")
    chart = BarChart()
    chart.title = "角色主观完成度"
    chart.y_axis.title = "完成度"
    chart.add_data(Reference(ws, min_col=3, min_row=role_start, max_row=role_start + len(ROLE_SUMMARY)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=role_start + 1, max_row=role_start + len(ROLE_SUMMARY)))
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "G10")
    row = add_section(ws, row, "全局进度")
    row = write_table(ws, row, ["模块", "状态", "完成度", "当前证据", "主要风险 / 缺口"], GLOBAL_PROGRESS, "GlobalProgress")
    row = add_section(ws, row, "本次已完成项")
    row = write_table(ws, row, ["日期", "完成项", "证据"], COMPLETED_ITEMS, "CompletedItems")
    row = add_section(ws, row, "待确认项")
    write_table(ws, row, ["编号", "待确认问题", "当前实现 / 建议"], PENDING_CONFIRMATIONS, "PendingConfirmations")

    for index, (role_name, role) in enumerate(ROLE_DATA.items(), 1):
        ws = workbook.create_sheet(role_name)
        setup_sheet(ws)
        add_title(ws, f"{role_name} 需求跟踪", role["desc"])
        row = 4
        row = add_section(ws, row, "主观进度")
        row = write_table(ws, row, ["区域", "状态", "完成度", "当前证据", "主要缺口"], role["summary"], f"Summary{index}")
        row = add_section(ws, row, "功能需求跟踪")
        row = write_table(ws, row, ["需求ID", "需求描述", "优先级", "状态", "完成度", "实现证据", "遇到的问题", "待说明/待确认", "备注"], role["requirements"], f"Req{index}")
        row = add_section(ws, row, "页面范围")
        write_table(ws, row, ["页面", "是否必需", "当前状态", "说明"], role["pages"], f"Pages{index}")

    style_workbook(workbook)
    workbook.save(OUT_FILE)


if __name__ == "__main__":
    build_workbook()
    print(OUT_FILE)
